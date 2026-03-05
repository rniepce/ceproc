"""
KPI Generator — Módulo 5: Gerador da Matriz de Indicadores e Metas.
===================================================================
Gera automaticamente o documento "PROPOSTA DE INDICADORES E METAS
PARA O PAINEL GERAL" em formato Excel (.xlsx) com 15 colunas.

Combina:
  1. Dados DETERMINÍSTICOS extraídos do parser BPMN (processo, subprocesso,
     produto/serviço, cliente, fontes de extração)
  2. Dados GENERATIVOS via LLM (indicadores, objetivo, metadados, fórmula,
     unidade, filtro, meta, periodicidade, polaridade)

Fallback: se a LLM falhar, gera o template formatado com colunas analíticas
em branco para preenchimento manual.
"""

import io
import json
import os
import re
import traceback
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

from bpmn_parser import BpmnModel


# ── Constants ───────────────────────────────────────────────────────

COLUMNS = [
    "N°",
    "PROCESSO",
    "SUBPROCESSO",
    "PRODUTO/SERVIÇO",
    "CLIENTE",
    "INDICADOR",
    "OBJETIVO",
    "METADADOS",
    "FONTE DE EXTRAÇÃO DOS METADADOS",
    "FÓRMULA DE CÁLCULO",
    "UNIDADE",
    "FILTRO",
    "META",
    "PERIODICIDADE",
    "POLARIDADE",
]

# Column widths (approximate)
COL_WIDTHS = {
    1: 5,    # N°
    2: 25,   # PROCESSO
    3: 22,   # SUBPROCESSO
    4: 25,   # PRODUTO/SERVIÇO
    5: 20,   # CLIENTE
    6: 30,   # INDICADOR
    7: 35,   # OBJETIVO
    8: 45,   # METADADOS
    9: 45,   # FONTE DE EXTRAÇÃO
    10: 50,  # FÓRMULA DE CÁLCULO
    11: 12,  # UNIDADE
    12: 25,  # FILTRO
    13: 15,  # META
    14: 15,  # PERIODICIDADE
    15: 18,  # POLARIDADE
}

# Wrap text columns (1-indexed)
WRAP_COLS = {8, 9, 10}  # METADADOS, FONTE DE EXTRAÇÃO, FÓRMULA

# Styles
TITLE_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
TITLE_FONT = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=9, bold=True, color="000000")
DATA_FONT = Font(name="Segoe UI", size=9)
FOOTER_FONT = Font(name="Segoe UI", size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ═══════════════════════════════════════════════════════════════════
# 1. EXTRACT DETERMINISTIC DATA FROM BPMN
# ═══════════════════════════════════════════════════════════════════

def _extract_bpmn_context(model: BpmnModel, cliente: str = "") -> dict:
    """
    Extract deterministic data from the parsed BPMN model.

    Returns dict with:
    - process_name: Main process name
    - subprocesses: List of subprocess names
    - products: List of end event names (product/service)
    - cliente: Client name
    - sources: List of data stores and data objects (systems/forms)
    - lanes: List of actor names
    """
    # Process name
    process_name = model.process_name or ""
    if not process_name and model.pools:
        process_name = model.pools[0].name

    # Subprocesses (look for subProcess tags)
    subprocesses = []
    for elem in model.elements.values():
        if elem.tag in ("subProcess", "callActivity"):
            subprocesses.append(elem.name or elem.id)

    # If no subprocesses, use milestones/intermediate events as phases
    if not subprocesses:
        for elem in model.elements.values():
            if elem.element_type == "event" and "intermediate" in (elem.event_subtype or ""):
                if elem.name:
                    subprocesses.append(elem.name)

    # Products (end events)
    products = []
    for elem in model.elements.values():
        if elem.element_type == "event" and "end" in (elem.event_subtype or ""):
            name = elem.name or ""
            if name and name.lower() not in ("fim", "final", "end"):
                products.append(name)
    if not products:
        products.append(f"{process_name} concluído" if process_name else "Processo concluído")

    # Sources (data stores + data objects)
    sources = []
    for obj in model.data_objects.values():
        if obj.name:
            prefix = "Sistema" if obj.data_type == "dataStore" else "Documento"
            sources.append(f"{prefix}: {obj.name}")

    # Lanes
    lanes = [lane.name for lane in model.lanes if lane.name]

    return {
        "process_name": process_name,
        "subprocesses": subprocesses,
        "products": products,
        "cliente": cliente,
        "sources": sources,
        "lanes": lanes,
    }


# ═══════════════════════════════════════════════════════════════════
# 2. LLM-GENERATED KPI PROPOSALS
# ═══════════════════════════════════════════════════════════════════

KPI_SYSTEM_PROMPT = """Você é um Analista de BI e Qualidade corporativo de um Tribunal de Justiça.
Com base no processo '{process_name}', nos atores '{lanes}' e nas fontes de dados
extraídas do diagrama '{sources}', proponha de 5 a 10 indicadores de desempenho (KPIs)
cobrindo Eficiência, Qualidade e Volume.

Retorne ESTRITAMENTE um array JSON contendo objetos com as seguintes chaves:
- indicador (Nome do KPI)
- objetivo (Objetivo do indicador)
- metadados (quais campos do sistema/formulário são necessários para o cálculo)
- formula_calculo (explique a lógica matemática de soma/contagem cruzando os formulários)
- unidade (ex: Unidade, Moeda, Dias, Percentual)
- filtro (quais colunas usar no dashboard)
- meta (valor meta sugerido, ex: ">= 90%", "<= 5 dias")
- periodicidade (ex: Mensal, Trimestral)
- polaridade (ex: Maior é Melhor, Menor é Melhor)

IMPORTANTE:
- Retorne APENAS o JSON válido (array de objetos), sem markdown, sem explicação.
- Os indicadores devem ser realistas e aplicáveis ao contexto do Tribunal de Justiça.
- As fórmulas devem referenciar os sistemas e documentos fornecidos quando possível.
"""


async def _generate_kpi_proposals(bpmn_context: dict) -> list:
    """
    Call LLM to generate KPI proposals.
    Returns list of dicts with KPI fields, or empty list on failure.
    """
    try:
        from gemini_engine import get_client, GPT_DEPLOYMENT

        client = get_client()

        process_name = bpmn_context.get("process_name", "Processo")
        lanes = ", ".join(bpmn_context.get("lanes", [])) or "Não identificados"
        sources = ", ".join(bpmn_context.get("sources", [])) or "Não identificados"

        system_prompt = KPI_SYSTEM_PROMPT.format(
            process_name=process_name,
            lanes=lanes,
            sources=sources,
        )

        user_prompt = f"""Processo: {process_name}

Atores/Lanes: {lanes}

Subprocessos/Fases: {', '.join(bpmn_context.get('subprocesses', [])) or 'Não identificados'}

Produtos/Saídas: {', '.join(bpmn_context.get('products', [])) or 'Não identificados'}

Fontes de Dados (Sistemas/Formulários): {sources}

Gere de 5 a 10 KPIs como array JSON."""

        response = client.chat.completions.create(
            model=GPT_DEPLOYMENT,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            max_completion_tokens=8000,
        )

        raw = response.choices[0].message.content.strip()

        # Clean markdown fences if present
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)

        kpis = json.loads(raw)
        if isinstance(kpis, list):
            return kpis

        return []

    except Exception as e:
        traceback.print_exc()
        print(f"[KPI Generator] LLM fallback — erro: {e}")
        return []


# ═══════════════════════════════════════════════════════════════════
# 3. BUILD DATAFRAME
# ═══════════════════════════════════════════════════════════════════

def _build_dataframe(bpmn_context: dict, kpi_proposals: list) -> pd.DataFrame:
    """
    Merge deterministic BPMN data with LLM-generated KPI proposals
    into a DataFrame with the 15-column schema.
    """
    process_name = bpmn_context.get("process_name", "")
    subprocesses = bpmn_context.get("subprocesses", [])
    products = bpmn_context.get("products", [])
    cliente = bpmn_context.get("cliente", "")
    sources = bpmn_context.get("sources", [])
    sources_str = "\n".join(sources) if sources else ""

    rows = []

    if kpi_proposals:
        for i, kpi in enumerate(kpi_proposals, 1):
            # Distribute subprocesses across KPIs if available
            subp = ""
            if subprocesses:
                subp = subprocesses[i % len(subprocesses)] if i <= len(subprocesses) else subprocesses[-1]

            # Distribute products
            prod = products[0] if products else ""
            if len(products) > 1 and i <= len(products):
                prod = products[i - 1]

            rows.append({
                "N°": i,
                "PROCESSO": process_name,
                "SUBPROCESSO": subp,
                "PRODUTO/SERVIÇO": prod,
                "CLIENTE": cliente,
                "INDICADOR": kpi.get("indicador", ""),
                "OBJETIVO": kpi.get("objetivo", ""),
                "METADADOS": kpi.get("metadados", ""),
                "FONTE DE EXTRAÇÃO DOS METADADOS": sources_str,
                "FÓRMULA DE CÁLCULO": kpi.get("formula_calculo", ""),
                "UNIDADE": kpi.get("unidade", ""),
                "FILTRO": kpi.get("filtro", ""),
                "META": kpi.get("meta", ""),
                "PERIODICIDADE": kpi.get("periodicidade", ""),
                "POLARIDADE": kpi.get("polaridade", ""),
            })
    else:
        # Fallback: generate 5 empty template rows
        for i in range(1, 6):
            subp = subprocesses[i - 1] if i <= len(subprocesses) else ""
            prod = products[0] if products else ""

            rows.append({
                "N°": i,
                "PROCESSO": process_name,
                "SUBPROCESSO": subp,
                "PRODUTO/SERVIÇO": prod,
                "CLIENTE": cliente,
                "INDICADOR": "",
                "OBJETIVO": "",
                "METADADOS": "",
                "FONTE DE EXTRAÇÃO DOS METADADOS": sources_str,
                "FÓRMULA DE CÁLCULO": "",
                "UNIDADE": "",
                "FILTRO": "",
                "META": "",
                "PERIODICIDADE": "",
                "POLARIDADE": "",
            })

    return pd.DataFrame(rows, columns=COLUMNS)


# ═══════════════════════════════════════════════════════════════════
# 4. EXCEL FORMATTING WITH OPENPYXL
# ═══════════════════════════════════════════════════════════════════

def _format_excel(wb, unidade: str, elaborado_por: str, aprovado_por: str):
    """
    Apply corporate formatting to the Excel workbook:
    - Row 1: merged title with dark background
    - Row 2: styled column headers
    - Data rows: borders, wrap text, fonts
    - Footer: elaborated/approved by
    - Freeze panes, column widths
    """
    ws = wb.active
    max_row = ws.max_row
    max_col = 15  # Always 15 columns

    # ── Insert title row at top ─────────────────────────────────
    ws.insert_rows(1)
    title_text = f"PROPOSTA DE INDICADORES E METAS PARA O PAINEL GERAL - {unidade}"
    ws.cell(row=1, column=1, value=title_text)

    # Merge A1:O1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    # Style title
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Style header row (now row 2) ────────────────────────────
    ws.row_dimensions[2].height = 30
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    # ── Style data rows ─────────────────────────────────────────
    # Data starts at row 3 (after title + header)
    data_end_row = max_row + 1  # +1 because we inserted title row

    for row in range(3, data_end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if col in WRAP_COLS:
                cell.alignment = Alignment(
                    vertical="top", wrap_text=True
                )
            elif col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="top")

    # ── Column widths ───────────────────────────────────────────
    for col_num, width in COL_WIDTHS.items():
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    # ── Freeze panes ────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Footer rows ─────────────────────────────────────────────
    today = datetime.now().strftime("%d/%m/%Y")

    footer_row_1 = data_end_row + 2
    footer_row_2 = data_end_row + 3

    # Row 1 footer
    ws.cell(row=footer_row_1, column=2, value=f"Elaborado por: {elaborado_por}").font = FOOTER_FONT
    ws.cell(row=footer_row_1, column=10, value=f"Data da elaboração: {today}").font = FOOTER_FONT

    # Row 2 footer
    ws.cell(row=footer_row_2, column=2, value=f"Aprovado por: {aprovado_por}").font = FOOTER_FONT
    ws.cell(row=footer_row_2, column=10, value=f"Data da aprovação: ___/___/______").font = FOOTER_FONT


# ═══════════════════════════════════════════════════════════════════
# 5. MAIN PUBLIC FUNCTION
# ═══════════════════════════════════════════════════════════════════

async def generate_kpi_excel(
    model: BpmnModel,
    unidade: str = "CEPROC",
    cliente: str = "",
    elaborado_por: str = "CEPROC",
    aprovado_por: str = "",
) -> bytes:
    """
    Generate the KPI Excel (.xlsx) file.

    Args:
        model: Parsed BPMN model from bpmn_parser.parse_bpmn()
        unidade: Unit name for the title (e.g., "CEPROC", "COGEPRO")
        cliente: Client name for deterministic column
        elaborado_por: Author name for footer
        aprovado_por: Approver name for footer

    Returns:
        Excel file as bytes (.xlsx)
    """
    # Step 1: Extract deterministic data
    bpmn_context = _extract_bpmn_context(model, cliente=cliente)

    # Step 2: Try to generate KPI proposals via LLM
    kpi_proposals = await _generate_kpi_proposals(bpmn_context)

    # Step 3: Build DataFrame
    df = _build_dataframe(bpmn_context, kpi_proposals)

    # Step 4: Write to Excel buffer (via pandas)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Indicadores", index=False, startrow=0)

    # Step 5: Reload and format with openpyxl
    buffer.seek(0)
    wb = load_workbook(buffer)
    _format_excel(wb, unidade, elaborado_por, aprovado_por)

    # Step 6: Save final
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
